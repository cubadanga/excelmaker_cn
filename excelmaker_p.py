
import pandas as pd
import openpyxl
from  openpyxl.styles  import  Alignment
from  openpyxl.styles.fonts  import  Font
import numpy as np
import random
import re
import os
import sys
import shutil
import time
from urllib.error import HTTPError
import urllib.request
from urllib.request import urlopen
from urllib.parse import urlparse, parse_qs
import configparser
from bs4 import BeautifulSoup
from colorama import init, Fore

#프린트문 색상 변경을 위해 초기화
np.set_printoptions(threshold=np.inf, linewidth=np.inf)
init()

print(Fore.LIGHTBLUE_EX + "你好？今天也度过愉快的一天吧。작성을 시작 합니다. 작성중..." )
print(Fore.RESET)

# ### 유저설정 시트와 상품정보 시트 추출
# * 엑셀에서 price 시트(입력시트) 추출
# * 엑셀에서 setting 시트 (셋팅시트) 추출
#중국버전에는 로그인 체크 하지 않음
'''
basedir = os.getcwd()
ini_dir = os.path.join(basedir,'set.ini')

# pc set.ini 파일의 저장된 pass워드 읽어오기
properties = configparser.ConfigParser()
properties.read(ini_dir)
password = properties['DEFAULT']['UserPass']

#웹에있는 password 텍스트 추출 함수
def getPtag(url):
    try:
        html = urlopen(url)
        
    except HTTPError as e:
        return None
    try:
        soup = BeautifulSoup(html,"html.parser")
        ptag = soup.find('p')
        
    except AttributeError as e:
        return None
    return ptag.text

#관리자 패스워드가 저장된 웹페이지 url을 전달하여 getPtag 함수 실행
passTag = getPtag("https://sites.google.com/view/test-exceldoc/pass")

def judge(password,passTag):
    if password == passTag:
        properties.set('DEFAULT','UserPass',password)
        with open('./set.ini','w',encoding='utf-8') as F:
            properties.write(F)
                
        print("이번 달 패스워드 체크 OK! 오늘도 파이팅!")
        pass
    else:
        print(Fore.RED + "오류 - 저장된 패스워드가 없거나 올바른 패스워드가 아닙니다. 패스워드는 단체방 금월 암호 공지를 확인하세요."+Fore.RESET+'\n')
        inputPass(password,passTag)

def inputPass(password,passTag):
    userpass = password
    passTag = passTag
    userpass = ""
    print('\n' + "패스워드를 입력해 주세요.")
    userPass = input()
    judge(userPass, passTag)

judge(password,passTag)
'''
try:
    df = pd.read_excel('./product.xlsx', sheet_name = 'write', header = 0)
    setpd = pd.read_excel('./product.xlsx', sheet_name = 'setting', header = 0)
    setpd = setpd.fillna('')
    
except ValueError as e:
    print(Fore.RED + '오류 - 엑셀 시트의 시트명이 다르거나 올바른 파일이 아닙니다.'+'\n')
    print(Fore.RESET + "엔터를 누르면 종료합니다.")
    aInput = input("")
    sys.exit()

except FileNotFoundError as e:
    print(Fore.RED + '오류 - product.xlsx 파일을 찾을 수 없습니다.'+'\n'+'이런 경우, 파일명이 잘못된 경우가 대부분이었습니다.'+' 이 파일은 필수 파일입니다.'+'\n')
    print(Fore.RESET + "엔터를 누르면 종료합니다.")
    aInput = input("")
    sys.exit()

#dfSourcing = pd.read_excel('./excel/sourcing/sourcing.xlsx', header = 0, index_col = 0)
pd.set_option('display.max_columns', None)

set_list = list(setpd['입력값'])
nickName = set_list[0]  #닉네임
as_info = set_list[1]   #as안내 내용
as_tel = set_list[2]    #A/s전화번호
factory_desc = set_list[3]  #제조사
brand_info = set_list[4]    #브랜드
discount_rate = float(set_list[5]) #표시 될 할인율
ship_method = set_list[6]   #배송비유형
qt_charge = set_list[7]     #수량별부과-수량
rship_price = set_list[8]   #기본배송비
check_method = set_list[9]  #배송비 결제방식
refund_ship = set_list[10]   #반품배송비
exchange_ship = set_list[11] #교환배송비
gift_desc = set_list[12] #사은품
point_tReview = set_list[13]    #텍스트리뷰 작성시 지급 포인트
point_photoReview = set_list[14]    #포토/동영상 리뷰 작성시 지급 포인트
point_monthText = set_list[15]  #한달사용 텍스트리뷰 작성시 지급 포인트
point_monthVideo = set_list[16] #한달사용 포토/동영상리뷰 작성시 지급 포인트
point_talktalk = set_list[17]   #톡톡친구/스토어찜고객 리뷰 작성시 지급 포인트
rate = float(set_list[18]) #환율
fomul = float(set_list[19])    #가격조정값
fee_naver = float(set_list[20])    #네이버수수료
marginMin = int(set_list[21])    #최소마진
naver_top = set_list[22]    #스스 상세페이지에 삽입되는 상단이미지
naver_bottom = set_list[23] #스스 상세페이지에 삽입되는 하단이미지
naver_bottom2 = set_list[24] #스스 상세페이지에 삽입되는 하단이미지 2
addDescBool = set_list[25]  #개인 상세페이지 상,하단 이미지 사용 유무

#계산이 필요한 금액은 숫자형으로 변경

# ### url 필드에서 상품ID 추출
shop_type =df['사이트'][0]
url_shop = df['url'][0]

def extract_id(site, url):
    parsed_url = urlparse(url)
    query_params = parse_qs(parsed_url.query)
    
    if site == 'taobao':
        product_id = query_params.get('id', [''])[0]
        product_url = "https://item.taobao.com/item.htm?id=" + product_id
        return product_id, product_url
    
    elif site == 'shop1688':
        file_name = os.path.splitext(os.path.basename(parsed_url.path))[0]
        product_id = file_name.split("_")[-1]
        product_url = "https://detail.1688.com/offer/" + product_id + ".html"
        return product_id, product_url
    
    else:
        return "", ""

productCord, product_url = extract_id(shop_type, url_shop)


print('추출한 id: ',productCord, '추출한 url',product_url)

if productCord =="":
    print(Fore.RED + '오류 - 입력한 주소가 해당 쇼핑몰의 주소인지 확인하세요. 추출된 코드가 없음.'+Fore.RESET+'\n')

else:
    print('상점타입: '+ shop_type)
    print('제품코드 추출 완료!: ' + productCord)
    
# 엑셀 기입용 제품코드
writePdCord = shop_type + '_' + productCord

# ### 상품명 추출
pName = df['상품명'][0]
print('제목 추출 완료!: ' + pName)

# 카테고리 번호 추출
categori = df['카테고리번호']
categori_list = list(categori)
categori_num = categori_list[0]

# 재고수량 추출
quanty = df['재고수량'].astype(str)
quanty_list = list(quanty)
quantyString = ",".join(quanty_list)

# 동영상 url 추출
#dfVurl = df.iloc[0,11:12]
videourl = str(df['동영상url'][0])

if videourl == 'nan':
    videourl = '동영상이 없습니다.'
    print('동영상 url은 없었습니다.')
    
else:
    print('동영상 url 복사완료!')

# ### 옵션명 제작
# * price 시트의 옵션명이 적혀있는 3개 행의 필드명을 추출하여 ','으로 구분하여 합친다.
# * price 시트의 행이 가변될 때 추출할 범위도 가변시켜 해당 옵션명을 받아온다.
# * 옵션의 조합이 몇개인지 판단하여 빈열 삭제

df_goods = df.iloc[0:,5:7]
df_goods.replace('', np.nan, inplace=True)
goods_Tclear = df_goods.dropna(axis=1)
colcount = len(goods_Tclear.columns)

gooddf = goods_Tclear.columns
optionTitle = str("\n".join(gooddf))

if colcount == 2:
    optionN1 = gooddf[0]
    optionN2 = gooddf[1]

elif colcount == 1:
    optionN1 = gooddf[0]

# 옵션에 관련된 데이터 열을 다 추출하여 계산에 사용함

df_optiongoods = df.iloc[0:,5:11]
df_optiongoods.replace('', np.nan, inplace=True)


goods_clear = df_optiongoods.dropna(axis=1).copy()

option_gooddf = goods_clear.columns
optionColcnt = len(goods_clear.columns)

optionT1 = []
optionT2 = []
optionT3 = []

if optionColcnt == 6:
    optionT1 = option_gooddf[0]
    optionT2 = option_gooddf[1]
    optionT3 = option_gooddf[2]


elif optionColcnt == 5:
    optionT1 = option_gooddf[0]
    optionT2 = option_gooddf[1]

elif optionColcnt == 4:
    optionT1 = option_gooddf[0]

# ### 기본 판매가 계산(옵션별 판매가격 계산)
# * 구매원가 = (상품가(상품가*수수료*환율)+배송비) prime_cost
# * 기본판매가 = 구매원가*가중치 price_min
# * 마진 = 기본가-스토어수수료-상품가-배송비 
# * 마진율 = 마진금액/기본가
goods_clear['구매원가'] = round(goods_clear['위안화']*1.03*int(rate)+goods_clear['실제배송비'],-2)
goods_clear['기본판매가'] = round(goods_clear['구매원가']*fomul,-2)
prime_cost = goods_clear['구매원가'].min()

# ==============================================#####
goods_clear['마진'] = round(goods_clear['기본판매가']-(goods_clear['기본판매가']*fee_naver/100)-goods_clear['구매원가'],-2)
goods_clear['마진율'] = round(goods_clear['마진']/goods_clear['기본판매가']*100,1)


# ### 옵션차액 계산
# * 기본판매가의 최소값, 최대값 추출
price_max = goods_clear['기본판매가'].max()
price_min = goods_clear['기본판매가'].min()


# ### 엑셀에 적힐 기본 판매가격 계산
# * 옵션별 판매가격이 차이가 없을 경우는 최소 금액이 판매가격이 됨
basePrice = np.int64(price_min)

# * 정해 놓은 마진 이상 남도록 최종판매가 다시 계산
# * setting시트에서 불러온 최소마진 설정값과 1차 계산 시 도출된 마진의 최소값과 비교한다.
# * 마진 리스트의 최소값이 < 최소마진(marginMin) 일 때 부족한 만큼 판매가격을 높여준다.

if marginMin > goods_clear['마진'].min():
    price_correction = round(((marginMin-goods_clear['마진'].min())*1.15),-2)
    price_correction = np.int64(round(price_correction,-2))

else :
    price_correction = 0
    price_correction = np.int64(round(price_correction,-2))
    


# * 최종판매가 = 기본판매가격+마진보정금액
tune_marginPrice = basePrice + price_correction

# 표시 판매가 계산
dp_price = round(tune_marginPrice / (1-discount_rate/100),-2)


goods_clear['옵션차액'] = round(goods_clear['기본판매가'] - price_min,-2)

#할인금액 계산
discount_price = dp_price - round(tune_marginPrice,-2)
discount_price = np.int64(discount_price)


# * 배송비 셋팅에서 유료 배송일 경우 판매가격에서 배송비를 차감하고 배송비 필드에 배송비 셋팅값을 입력한다.
if ship_method == "유료":
    finalPrice = dp_price-rship_price
    finalPrice = np.int64(round(finalPrice,-2))

else:
    finalPrice = dp_price
    finalPrice = np.int64(round(finalPrice,-2))

print('가격계산 완료!')

tuneMargin = round(tune_marginPrice-goods_clear['구매원가'].min()-(tune_marginPrice*fee_naver/100),-2)
tuneMarginRate = round(tuneMargin/tune_marginPrice*100,0)

# ### 옵션항목 뽑기
option_list1 = []
option_list2 = []
option_list3 = []
opPrice_list = []

df_gc = goods_clear.astype(str)

#optionT1, T2, T3는 옵션항목의 필드명들 순서대로 표기
#입력가격이 몇 번째 옵션을 기준하여 작성되었는지 판정
#옵션1, 옵션2 칸에 입력한 옵션내용 데이터를 중복제거하고 데이터프레임을 필터링해본다.
#필터링된 데이터프레임에서 옵션차액 필드 내의 데이터의 중복 검사를 해본다.
#중복검사에서 1개가 나오면 모든 데이터가 1개의 가격으로 쓰여져 있다는 뜻..그러므로 주요한 가격이 아님
#2개 이상이 나오면 그 필터링된 제목을 가진 옵션이 가격을 결정하는 것
#그 제목의 순서대로 옵션차액을 기록한다.

deff_price = ""
optionPrice = ""
deff_list = []
zerodeff_list = []

if optionColcnt == 5:
    df_gcprice = df_gc.drop_duplicates(subset=optionT1,ignore_index=False)
    df_subset1 = df_gcprice['옵션차액'].drop_duplicates()
    dupPriceCnt1 = df_subset1.value_counts().sum(axis=0)

    df_gcprice2 = df_gc.drop_duplicates(subset=optionT2,ignore_index=False)
    df_subset2 = df_gcprice2['옵션차액'].drop_duplicates()
    dupPriceCnt2 = df_subset2.value_counts().sum(axis=0)

    if dupPriceCnt1 >= 2:
        print('"첫번째 옵션이 가격을 결정합니다."')
        df_option1 = df_gc[optionT1].drop_duplicates()
                
        for op in df_option1:
            option_deff = goods_clear.loc[goods_clear[optionT1] == op]
            intdeff = option_deff['옵션차액'].drop_duplicates()

            strdeff = np.int64(intdeff.min())
            deff_list.append(strdeff)

        for i in range(len(deff_list)):
            deff_list[i] = str(deff_list[i])
            zerodeff_list.append("0")

        deff_price = str(",".join(deff_list))
        zero_deff = str(",".join(zerodeff_list))
        optionPrice = deff_price + '\n' + zero_deff # optionPrice

    elif dupPriceCnt2 >= 2:
        print('"두번째 옵션이 가격을 결정합니다."')
        df_option1 = df_gc[optionT2].drop_duplicates()
        
        for op in df_option1:
            option_deff = goods_clear.loc[goods_clear[optionT2] == op]
            intdeff = option_deff['옵션차액'].drop_duplicates()
            strdeff = np.int64(intdeff.min())
            deff_list.append(strdeff)

        for i in range(len(deff_list)):
            deff_list[i] = str(deff_list[i])
            zerodeff_list.append("0")

        deff_price = str(",".join(deff_list))
        zero_deff = str(",".join(zerodeff_list))
        optionPrice = zero_deff+'\n'+ deff_price

    else:
        print('옵션의 가격이 모두 동일합니다.')

elif optionColcnt == 4:
    df_option1 = df_gc[optionT1].drop_duplicates()  # 첫번째 필드의 데이터들을 프레임에 담는다.

    # 일단 같은 옵션명과 금액을 가진 놈들을 뽑아 중복제거 후 리스트에 담는다.
    # int로 변경 후 다른 이름을 또 검색해서 중복제거 후 계속 추가한다.
    # 완성된 리스트를 스트링으로 변환한다.
    for op in df_option1:
        option_deff = goods_clear.loc[goods_clear[optionT1] == op]
        intdeff = option_deff['옵션차액'].drop_duplicates()
        strdeff = np.int64(intdeff.min())
        deff_list.append(strdeff)
    # join 함수를 사용할 때는 리스트 내의 인자들이 모두 string 형태여야 한다. 그러니깐.
    # 구성요소 들을 for문으로 돌면서 스트링으로 바꿔준 다음 join으로 합친다.
    for i in range(len(deff_list)):
        deff_list[i] = str(deff_list[i])
    deff_price = str(",".join(deff_list))
    optionPrice = deff_price

# 네이버가 요구하는 양식으로 데이터를 편집하여 스트링으로 저장

print('옵션 작성 완료!')

if optionColcnt == 6:
    df_option1 = goods_clear[optionT1].drop_duplicates()
    df_option2 = goods_clear[optionT2].drop_duplicates()
    df_option3 = goods_clear[optionT3].drop_duplicates()
    list_option1 = df_option1.values.tolist()#담겨진 데이터들 중 중복 삭제하고 유일한 값들만 모아서 프레임에 저장
    list_option2 = df_option2.values.tolist()
    list_option3 = df_option3.values.tolist()

    optionDesc1 = ",".join(map(str,list_option1))
    optionDesc2 = ",".join(map(str,list_option2))
    optionDesc3 = ",".join(map(str,list_option3))
    optionValue = optionDesc1 + '\n' + optionDesc2 + '\n' + optionDesc3

elif optionColcnt == 5:
    df_option1 = goods_clear[optionT1].drop_duplicates()
    df_option2 = goods_clear[optionT2].drop_duplicates()
    list_option1 = df_option1.values.tolist()  # 담겨진 데이터들 중 중복 삭제하고 유일한 값들만 모아서 프레임에 저장
    list_option2 = df_option2.values.tolist()
    optionDesc1 = ",".join(map(str,list_option1))
    optionDesc2 = ",".join(map(str,list_option2))
    optionValue = optionDesc1 + '\n' + optionDesc2

elif optionColcnt == 4:
    df_option1 = goods_clear[optionT1].drop_duplicates()
    list_option1 = df_option1.values.tolist()
    optionDesc1 = ",".join(map(str,list_option1))
    optionValue = optionDesc1
    optionPrice = deff_price  # optionPrice
    txtOption1 = df_gc[optionT1].drop_duplicates()
    df_OpDescTitle = txtOption1

#상세페이지 작성
try:
    dpHtml = df['상세페이지']
    dpHtml_list = list(dpHtml)
    preDescPages = dpHtml_list[0]
    descPages2 = re.sub("img referrerpolicy='no-referrer'|{LINK}|", "", preDescPages)
    descPages1 = re.sub("< ", "<", descPages2)+'\n'
    descPages = '<div align="center">' + descPages1 + '</div>'

except TypeError:
    print(Fore.RED + '오류 - product.xlsx->상세페이지 필드에 url이 없거나 잘못 되었습니다.')
    print(Fore.RESET + "엔터를 누르면 종료합니다.")
    aInput = input("")
    sys.exit()
    

descPname = '<br><br><h1 style="text-align: center;"><strong>' + pName + "</strong></h1><br><br>"+'\n'
naverTop = '<div align="center"><img src="' + naver_top + '"/></div>'+'\n'
naverBottom = '<div align="center"><img src="' + naver_bottom + '"/></div>'+'\n'
naverBottom2 = '<div align="center"><img src="' + naver_bottom2 + '"/></div>'+'\n'
#shop11Top = '<img src="' + shop11st_top + '"/>'+'\n'
#shop11stBottom = '<img src="' + shop11st_bottom + '"/>'+'\n'

try:
    df_opurl = df.iloc[0:,4:6]
    df_filter = df_opurl.drop_duplicates(subset=optionT1,ignore_index=False)
    img_option = df_filter['옵션이미지']
    img_optionTag = img_option.str.replace('<img src="','')
    img_optionTag = img_optionTag.str.replace("<img src='",'')
    img_optionTag = img_optionTag.str.replace('"/>','')
    img_optionTag = img_optionTag.str.replace("'/>",'')
    img_optionTag = img_optionTag.str.replace('" />','')
    img_optionTag = img_optionTag.str.replace("' />",'')
    op_imgurls = img_optionTag.values.tolist()
except KeyError:
    print(Fore.RED + '오류 - 옵션이미지 필드에 url이 없거나 잘못 되었습니다.')
    print(Fore.RESET + "엔터를 누르면 종료합니다.")
    aInput = input("")
    sys.exit()
    
OpTitle = df_filter[optionT1]
op_titlelist = OpTitle.values.tolist()
optionLen = len(op_titlelist)

opjoin_list = []
cntj=1

for i in range(optionLen):
    
    try :
        strtitle = '<div align="center"><div><h2><strong>옵션'+str(cntj)+'. '+ op_titlelist[i]+'</strong></h2></div>'
        strImg = '<div align="center"><img src="'+op_imgurls[i]+'"/></div><br><br>'
        opjoin_list.append(strtitle+strImg)
        cntj += 1
    
    except TypeError as e:
        print(Fore.RED + '오류 - 옵션 url을 입력하지 않은 것 같습니다.')
        print(Fore.RESET + "엔터를 누르면 종료합니다.")
        aInput = input("")
        sys.exit()
    
opjoinStr = str("\n".join(opjoin_list))
optionHtml = '<br><div align="center"><img src="https://i.ibb.co/vZpWH4Z/option-Img.png" alt="option-Img" border="0"></div><br>'+ opjoinStr

descNaver = ""
desc11st = ""
p_desc = ""

if addDescBool == 0:
    descNaver = naverTop + descPname + descPages + optionHtml + naverBottom + naverBottom2
    #desc11st = shop11Top + descPname + descPages + optionHtml + shop11stBottom
    p_desc = descPname + descPages + optionHtml
    descNaver = descNaver.replace('<img src=""/>', '')
    descPN = "<div align='center'>" + descNaver + "</div>"

elif addDescBool ==1:
    descNaver = descPname + descPages + optionHtml
    #desc11st = descPname + descPages + optionHtml
    p_desc = descPname + descPages + optionHtml
    descNaver = descNaver.replace('<img src=""/>', '')
    descPN = "<div align='center'>" + descNaver + "</div>"

else:
    print(Fore.RED + "오류 - 상하단 이미지 등록 여부가 잘못 입력 되었습니다." + Fore.RESET+'\n')


print("상세페이지 작성 완료!")

# ### 엑셀에 기재될 배송비
if ship_method == "유료":
    ship_price = rship_price

else:
    ship_price = 0

ship_price = str(ship_price)

# ###이미지 파일을 불러옴
file_path = './mainImages'
output_path = './mainImages'
try:
    file_names = os.listdir(file_path)

except FileNotFoundError as e:
    print(Fore.RED + '오류 - mainImage(메인이미지) 폴더가 존재하지 않습니다.')
    print(Fore.RESET + "엔터를 누르면 종료합니다.")
    aInput = input("")
    sys.exit()

if len(file_names) > 0:
    i = 1
    j = 0
    images=[]

    for name in file_names:
        src = os.path.join(file_path,name)
        dst = productCord + '-' + str(i) + '.jpg'
        images.append(dst)
        dst = os.path.join(output_path,dst)
        os.rename(src,dst)
        i += 1
        j += 1

    mainImage = images[0]
    del images[0]
    subImages = ",".join(images)
    #destination = "./excel/images/" + productCord + '-' + str(i) + '.jpg'

    def folder_file_copy():
        
        file_dir = os.path.dirname('./mainImages/')
        file_cnt = 1
        for path, dirs, files in os.walk(file_dir):
            for file in files:
                file_path = os.path.join(path,file)
                file_cnt += 1
                dest_path = './excel/' + file
                shutil.copy(file_path, dest_path)
                
    folder_file_copy()

else:
    print(Fore.RED + "오류 - 메인이미지 폴더에 이미지가 없습니다.")
    mainImage = ""
    subImages = ""
print(Fore.RESET + "메인이미지 수정/이동 완료!")

#스마트스토어 필드명 불러오기

store_field = pd.read_excel('./product.xlsx', sheet_name = 'store', header = 0)
storeField_list = list(store_field['네이버'])

#스마트스토어 본인용 엑셀파일 생성
wb = openpyxl.Workbook()
ws = wb.active
ws.append(storeField_list)

#카테고리 불러오기
ncategori = pd.read_excel('./product.xlsx', sheet_name = 'categori_naver', header = 0)
catStr = int(categori_num)
df_cat = ncategori.loc[ncategori['카테고리번호'] == catStr].fillna("")

strCalevel1 = df_cat['대분류'].to_string(index=False)
strCalevel2 = df_cat['중분류'].to_string(index=False)
strCalevel3 = df_cat['소분류'].to_string(index=False)
strCalevel4 = df_cat['세분류'].to_string(index=False)

tday = time.time()
tday_s = time.strftime('%Y%m%d-%H%M%S',time.localtime(time.time()))
tday_f = time.strftime('%Y-%m-%d',time.localtime(time.time()))

ws["A2"].value = "신상품"
ws["B2"].value = categori_num
ws["C2"].value = pName
ws["D2"].value = dp_price
ws["E2"].value = "999"
ws["F2"].value = as_info
ws["G2"].value = as_tel
ws["H2"].value = mainImage
ws["I2"].value = subImages
ws["J2"].value = descNaver
ws["k2"].value = writePdCord
ws["L2"].value = " "
ws["M2"].value = factory_desc
ws["N2"].value = brand_info
ws["O2"].value = " "
ws["P2"].value = " "
ws["Q2"].value = "과세상품"
ws["R2"].value = "Y"
ws["S2"].value = "Y"
ws["T2"].number_format = '"0"#'

cellFormat = ws["T2"]
cellFormat.number_format = '@'
ws["T2"].value = "0200037"

ws["U2"].value = factory_desc
ws["V2"].value = "N"
ws["W2"].value = " "
ws["X2"].value = "택배, 소포, 등기"
ws["Y2"].value = "CJGLS"
ws["Z2"].value = ship_method
ws["AA2"].value = rship_price
ws["AB2"].value = check_method
ws["AC2"].value = " "
ws["AD2"].value = qt_charge
ws["AE2"].value = refund_ship
ws["AF2"].value = exchange_ship
ws["AG2"].value = " "
ws["AH2"].value = " "
ws["AI2"].value = " "
ws["AJ2"].value = discount_price
ws["AK2"].value = "원"
ws["AL2"].value = " "
ws["AM2"].value = " "
ws["AN2"].value = " "
ws["AO2"].value = " "
ws["AP2"].value = " "
ws["AQ2"].value = " "
ws["AR2"].value = point_tReview
ws["AS2"].value = point_photoReview
ws["AT2"].value = point_monthText
ws["AU2"].value = point_monthVideo
ws["AV2"].value = point_talktalk
#ws["AW2"].value = " "
ws["AX2"].value = gift_desc
ws["AY2"].value = "조합형"
ws["AZ2"].value = optionTitle
ws["BA2"].value = optionValue
ws["BB2"].value = optionPrice
ws["BC2"].value = quantyString
ws["BD2"].value = " "
ws["BE2"].value = " "
ws["BF2"].value = " "
ws["BG2"].value = " "
ws["BH2"].value = "상세페이지 참조"
ws["BI2"].value = "상세페이지 참조"
ws["BJ2"].value = "상세페이지 참조"
ws["BK2"].value = "상세페이지 참조"
ws["BL2"].value = "N"
ws["BM2"].value = " "
ws["BN2"].value = " "
ws["BO2"].value = " "
ws["BP2"].value = " "
ws["BQ2"].value = " "
ws["BR2"].value = " "
ws["BS2"].value = " "
ws["BT2"].value = " "
ws["BU2"].value = " "
ws["BV2"].value = nickName # 작성자
ws["BW2"].value = tday_f # 소싱일
ws["BX2"].value = writePdCord
ws["BY2"].value = pName
ws["BZ2"].value = product_url
ws["CA2"].value = goods_clear['위안화'].min()
ws["CB2"].value =rate
ws["CC2"].value = goods_clear['실제배송비'].min()
ws["CD2"].value = round(prime_cost,-2)
ws["CE2"].value = round(tune_marginPrice,-2)
ws["CF2"].value = round(tuneMargin,1)
ws["CG2"].value = round(tuneMarginRate,1)
ws["CH2"].value = fomul
ws["CI2"].value = marginMin
ws["CJ2"].value = categori_num
ws["Ck2"].value = strCalevel1
ws["CL2"].value = strCalevel2
ws["CM2"].value = strCalevel3
ws["CN2"].value = strCalevel4

new_fileName = ('./excel/'+productCord+'_'+'개인용'+'_'+tday_s+'.xlsx')
wb.save(new_fileName)
print("개인용파일 작성완료!")

store_field2 = pd.read_excel('./product.xlsx', sheet_name = 'store', header = 0)
storeField_list2 = list(store_field2['네이버'])

#스마트스토어 배포용 엑셀파일 생성
p_wb = openpyxl.Workbook()
p_ws = p_wb.active
p_ws.append(storeField_list2)
p_ws["A2"].value = "신상품"
p_ws["B2"].value = categori_num
p_ws["C2"].value = pName
p_ws["D2"].value = dp_price
p_ws["E2"].value = "999"
p_ws["F2"].value = "as_info"
p_ws["G2"].value = "000-000-0000"
p_ws["H2"].value = mainImage
p_ws["I2"].value = subImages
p_ws["J2"].value = descPN
p_ws["k2"].value = writePdCord
p_ws["L2"].value = " "
p_ws["M2"].value = "factory_desc"
p_ws["N2"].value = "brand_info"
p_ws["O2"].value = " "
p_ws["P2"].value = " "
p_ws["Q2"].value = "과세상품"
p_ws["R2"].value = "Y"
p_ws["S2"].value = "Y"

cellFormat = p_ws["T2"]
cellFormat.number_format = '@'

p_ws["T2"].value = "0200037"
p_ws["U2"].value = "factory_desc"
p_ws["V2"].value = "N"
p_ws["W2"].value = " "
p_ws["X2"].value = "택배, 소포, 등기"
p_ws["Y2"].value = "CJGLS"
p_ws["Z2"].value = ship_method
p_ws["AA2"].value = rship_price
p_ws["AB2"].value = check_method
p_ws["AC2"].value = " "
p_ws["AD2"].value = qt_charge
p_ws["AE2"].value = refund_ship
p_ws["AF2"].value = exchange_ship
p_ws["AG2"].value = " "
p_ws["AH2"].value = " "
p_ws["AI2"].value = " "
p_ws["AJ2"].value = discount_price
p_ws["AK2"].value = "원"
p_ws["AL2"].value = " "
p_ws["AM2"].value = " "
p_ws["AN2"].value = " "
p_ws["AO2"].value = " "
p_ws["AP2"].value = " "
p_ws["AQ2"].value = " "
p_ws["AR2"].value = point_tReview
p_ws["AS2"].value = point_photoReview
p_ws["AT2"].value = point_monthText
p_ws["AU2"].value = point_monthVideo
p_ws["AV2"].value = point_talktalk
p_ws["AW2"].value = " "
p_ws["AX2"].value = gift_desc
p_ws["AY2"].value = "조합형"
p_ws["AZ2"].value = optionTitle
p_ws["BA2"].value = optionValue
p_ws["BB2"].value = optionPrice
p_ws["BC2"].value = quantyString
p_ws["BD2"].value = " "
p_ws["BE2"].value = " "
p_ws["BF2"].value = " "
p_ws["BG2"].value = " "
p_ws["BH2"].value = "상세페이지 참조"
p_ws["BI2"].value = "상세페이지 참조"
p_ws["BJ2"].value = "상세페이지 참조"
p_ws["BK2"].value = "상세페이지 참조"
p_ws["BL2"].value = "N"
p_ws["BM2"].value = " "
p_ws["BN2"].value = " "
p_ws["BO2"].value = " "
p_ws["BP2"].value = " "
p_ws["BQ2"].value = " "
p_ws["BR2"].value = " "
p_ws["BS2"].value = " "
p_ws["BT2"].value = " "
p_ws["BU2"].value = " "
p_ws["BV2"].value = nickName # 작성자
p_ws["BW2"].value = tday_f # 소싱일
p_ws["BX2"].value = writePdCord
p_ws["BY2"].value = pName
p_ws["BZ2"].value = product_url
p_ws["CA2"].value = goods_clear['위안화'].min()
p_ws["CB2"].value =rate
p_ws["CC2"].value = goods_clear['실제배송비'].min()
p_ws["CD2"].value = round(prime_cost,-2)
p_ws["CE2"].value = round(tune_marginPrice,-2)
p_ws["CF2"].value = round(tuneMargin,1)
p_ws["CG2"].value = round(tuneMarginRate,1)
p_ws["CH2"].value = fomul
p_ws["CI2"].value = marginMin
p_ws["CJ2"].value = categori_num
p_ws["Ck2"].value = strCalevel1
p_ws["CL2"].value = strCalevel2
p_ws["CM2"].value = strCalevel3
p_ws["CN2"].value = strCalevel4


new_fileName = ('./excel/'+productCord+'_'+'배포용'+'_'+tday_s+'.xlsx')
p_wb.save(new_fileName)
print("배포용파일 작성완료!")

# 이미지 저장용 폴더 생성(중국버전에서는 제외)

tday = time.time()
fday = time.strftime('%Y%m%d',time.localtime(time.time()))

def createFolder(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        print (Fore.RED + '오류 - Creating directory. ' +  directory)
        print(Fore.RESET + "엔터를 누르면 종료합니다.")
        aInput = input("")
        sys.exit()

#pathf = ""
pathf = './excel/'+ productCord
#pathDesc = './excel/'+ productCord +'/Desc'
#pathOption = './excel/'+ productCord +'/Option'
pathBackup = './excel/product_backup'
createFolder(pathf)
#createFolder(pathDesc)
#createFolder(pathOption)
createFolder(pathBackup)
print('폴더 생성 완료!'+'\n')

# 옵션 이미지 다운로드(중국버전에서는 다운하지 않음)
'''
optionNum = 0
###
try:
    for i in op_imgurls: 
        file_ext = i.split('.')[-1] # 확장자 추출
        path = pathOption + '/' + productCord + '_option_' + str(optionNum)+'.' + file_ext
        random_number = round(random.uniform(0.07, 0.2), 2)
        
        time.sleep(random_number)
        urllib.request.urlretrieve(i, path)
        print(str(optionNum)+'번 옵션 이미지 다운로드 성공')
        optionNum +=1
        
    # 상세 이미지 다운로드
    descimgNum = 0
    descPages = descPages.replace('?getAvatar=avatar','')
    modUrls1 = re.findall('<img.*?src="(.*?)".*?>', descPages)
    modUrls2 = re.findall("<img.*?src='(.*?)'.*?>", descPages)
    modUrls = modUrls1 + modUrls2

except urllib.error.HTTPError:
    print(Fore.RED + '오류 - 크롬 브라우저로 타오바오에 로그인이 필요하거나 올바른 옴션 url이 아닙니다.')
    print(Fore.RESET + "엔터를 누르면 종료합니다.")
    aInput = input("")
    sys.exit()


try:    
    for i in modUrls: 
        file_ext = i.split('.')[-1] # 확장자 추출
        path = pathDesc + '/' + productCord + '_desc_' + str(descimgNum)+'.' + file_ext
        random_number = round(random.uniform(0.02, 0.2), 2)
        
        time.sleep(random_number)
        urllib.request.urlretrieve(i, path)
        print(str(descimgNum)+'번 상세 이미지 다운로드 성공')
        descimgNum +=1

except urllib.error.HTTPError:
    print(Fore.RED + '오류 - 해외쇼핑몰 로그인이 필요하거나 올바른 상세 url이 아닙니다.')
    print(Fore.RESET + "엔터를 누르면 종료합니다.")
    aInput = input("")
    sys.exit()
'''
fVideoUrl = open(pathf+'/동영상주소.txt','w')
#fVideoUrl = open('./excel/' + productCord + '/동영상주소.txt','w')<--한국버전용
fVideoUrl.write(videourl)    
fVideoUrl.close()

copy_df = df
copy_df = df.to_excel(excel_writer=pathBackup+'/product_'+tday_s+'.xlsx', index=False)
print('\n'+ Fore.LIGHTBLUE_EX + "완성! 엔터를 누르면 종료합니다." + Fore.RESET)
aInput = input("")